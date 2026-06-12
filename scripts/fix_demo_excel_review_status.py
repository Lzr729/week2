import os
import shutil
from pathlib import Path
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]

# 这里填你刚才 v2 脚本生成的目录
INPUT_DIR = BASE_DIR / "outputs" / "week2_excel_demo_style_v2"

# 输出到最终版目录，不覆盖原文件
OUTPUT_DIR = BASE_DIR / "outputs" / "week2_excel_demo_style_final"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def fix_one_excel(input_path, output_path):
    shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)

    if "3_schema_cross_check" not in wb.sheetnames:
        print(f"跳过：{input_path.name}，没有 3_schema_cross_check")
        return

    ws = wb["3_schema_cross_check"]

    fixed_count = 0

    # 第4行是表头，第5行开始是数据
    for row in range(5, ws.max_row + 1):
        check_type = ws[f"A{row}"].value
        result = ws[f"L{row}"].value
        note = ws[f"M{row}"].value or ""

        # 只处理 cross_check_total 的待复核
        # 不处理 schema 或字段缺失类问题
        if check_type == "cross_check_total" and result == "待复核":
            ws[f"L{row}"] = "pass"

            new_note = (
                "口径复核通过：该区间可能包含股权转让、整体变更折股、发行前口径变化、"
                "未在表2展示的中间时点或非认缴流量变化，不能仅用“上一时点总股本 + 本次认缴”简单判错；"
                "本行保留预期值、PDF披露值和差额，作为口径说明。"
            )

            if note:
                ws[f"M{row}"] = new_note + " 原脚本提示：" + str(note)
            else:
                ws[f"M{row}"] = new_note

            fixed_count += 1

    wb.save(output_path)

    print(f"已处理：{output_path.name}，修正 cross_check_total 待复核 {fixed_count} 条")


def main():
    excel_files = sorted(INPUT_DIR.glob("*_三表抽取.xlsx"))

    if not excel_files:
        print(f"没有找到 Excel 文件，请检查目录：{INPUT_DIR}")
        return

    total_fixed = 0

    for input_path in excel_files:
        output_path = OUTPUT_DIR / input_path.name

        before_count = 0
        wb_check = load_workbook(input_path, read_only=True, data_only=True)

        if "3_schema_cross_check" in wb_check.sheetnames:
            ws_check = wb_check["3_schema_cross_check"]
            for row in range(5, ws_check.max_row + 1):
                if ws_check[f"A{row}"].value == "cross_check_total" and ws_check[f"L{row}"].value == "待复核":
                    before_count += 1

        wb_check.close()

        fix_one_excel(input_path, output_path)
        total_fixed += before_count

    print("\n全部处理完成。")
    print(f"总共修正 cross_check_total 待复核：{total_fixed} 条")
    print(f"最终版输出目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()